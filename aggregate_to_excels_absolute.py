"""
Aggregate Benchmark Results to Excel (Absolute Scoring Version)
Converts JSONL and CSV results into organized Excel workbooks.
"""

import argparse
import json
import csv
import os
from typing import List, Dict
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def load_inference_results(jsonl_file: str) -> List[Dict]:
    """Load inference results from JSONL."""
    results = []

    with open(jsonl_file, 'r', encoding='utf-8') as f:
        for line in f:
            if line.strip():
                results.append(json.loads(line))

    print(f"Loaded {len(results)} inference results")
    return results


def load_judge_scores(csv_file: str) -> List[Dict]:
    """Load absolute judge scores from CSV."""
    scores = []

    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Convert numeric fields
            row['factuality'] = float(row['factuality'])
            row['helpfulness'] = float(row['helpfulness'])
            row['structure'] = float(row['structure'])
            row['conciseness'] = float(row['conciseness'])
            row['total_score'] = float(row['total_score'])
            row['question_id'] = int(row['question_id'])

            scores.append(row)

    print(f"Loaded {len(scores)} judge scores")
    return scores


def organize_by_model(data: List[Dict], key: str = 'model_id') -> Dict[str, List[Dict]]:
    """Organize data by model_id."""
    organized = defaultdict(list)

    for item in data:
        model_id = item[key]
        organized[model_id].append(item)

    return dict(organized)


def create_quality_excel(scores: List[Dict], output_file: str):
    """Create Excel workbook with quality scores per model."""
    print(f"\nCreating quality scores workbook: {output_file}")

    # Organize scores by model
    model_scores = organize_by_model(scores)

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create sheet for each model
    for model_id in sorted(model_scores.keys()):
        ws = wb.create_sheet(title=model_id[:31])  # Excel sheet name limit

        # Headers
        headers = [
            'question_id', 'category', 'level',
            'Factuality', 'Helpfulness', 'Structure', 'Conciseness',
            'Total Score', 'Reasoning'
        ]
        ws.append(headers)

        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add data rows
        for score in sorted(model_scores[model_id], key=lambda x: (x['question_id'], x['level'])):
            ws.append([
                score['question_id'],
                score['category'],
                score['level'],
                score['factuality'],
                score['helpfulness'],
                score['structure'],
                score['conciseness'],
                score['total_score'],
                score['reasoning'][:100] if len(score['reasoning']) > 100 else score['reasoning']
            ])

        # Calculate level averages
        ws.append([])  # Empty row
        ws.append(['Level Averages'])

        level_stats = defaultdict(lambda: {
            'factuality': [], 'helpfulness': [], 'structure': [],
            'conciseness': [], 'total': []
        })

        for score in model_scores[model_id]:
            level = score['level']
            level_stats[level]['factuality'].append(score['factuality'])
            level_stats[level]['helpfulness'].append(score['helpfulness'])
            level_stats[level]['structure'].append(score['structure'])
            level_stats[level]['conciseness'].append(score['conciseness'])
            level_stats[level]['total'].append(score['total_score'])

        # Write averages
        for level in ['L0', 'L1', 'L2', 'L3', 'P']:
            if level in level_stats and level_stats[level]['total']:
                avg_fact = sum(level_stats[level]['factuality']) / len(level_stats[level]['factuality'])
                avg_help = sum(level_stats[level]['helpfulness']) / len(level_stats[level]['helpfulness'])
                avg_struct = sum(level_stats[level]['structure']) / len(level_stats[level]['structure'])
                avg_conc = sum(level_stats[level]['conciseness']) / len(level_stats[level]['conciseness'])
                avg_total = sum(level_stats[level]['total']) / len(level_stats[level]['total'])

                ws.append([
                    '',
                    'Average',
                    level,
                    round(avg_fact, 2),
                    round(avg_help, 2),
                    round(avg_struct, 2),
                    round(avg_conc, 2),
                    round(avg_total, 2),
                    ''
                ])

        # Format columns
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 40

        # Enable text wrapping for reasoning
        for row in ws.iter_rows(min_row=2):
            row[8].alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(output_file)
    print(f"  Saved: {output_file}")


def create_metric_excel(results: List[Dict], metric_name: str, metric_key: str, output_file: str):
    """
    Create Excel workbook for a specific metric (energy, latency, output_tokens).

    Args:
        results: List of inference results
        metric_name: Display name (e.g., "Energy")
        metric_key: Key in result dict (e.g., "energy_joule")
        output_file: Output file path
    """
    print(f"\nCreating {metric_name} workbook: {output_file}")

    # Organize by model
    model_results = organize_by_model(results)

    wb = Workbook()
    wb.remove(wb.active)

    # Create sheet for each model
    for model_id in sorted(model_results.keys()):
        ws = wb.create_sheet(title=model_id[:31])

        # Get unique questions and levels
        questions = sorted(set(r['question_id'] for r in model_results[model_id]))
        levels = ['L0', 'L1', 'L2', 'L3', 'P']

        # Create headers: question_id, category, L0, L1, L2, L3, P
        headers = ['question_id', 'category'] + levels
        ws.append(headers)

        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Organize data by question and level
        data_matrix = {}

        for result in model_results[model_id]:
            qid = result['question_id']
            level = result['level']
            category = result['category']
            value = result.get(metric_key, 0)

            if qid not in data_matrix:
                data_matrix[qid] = {'category': category}

            data_matrix[qid][level] = value

        # Add data rows
        for qid in questions:
            if qid not in data_matrix:
                continue

            row = [qid, data_matrix[qid].get('category', '')]

            for level in levels:
                value = data_matrix[qid].get(level, '')
                if isinstance(value, (int, float)):
                    row.append(round(value, 3))
                else:
                    row.append(value)

            ws.append(row)

        # Add summary statistics at the bottom
        ws.append([])  # Empty row
        ws.append(['Summary Statistics'])

        summary_row_start = ws.max_row

        # Calculate means for each level
        means = ['Mean', '']

        for level in levels:
            values = [data_matrix[qid].get(level, 0) for qid in questions if level in data_matrix[qid]]
            if values:
                mean_val = sum(values) / len(values)
                means.append(round(mean_val, 3))
            else:
                means.append(0)

        ws.append(means)

        # Style summary
        for cell in ws[summary_row_start]:
            cell.font = Font(bold=True)

        # Format columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15

        for i, level in enumerate(levels, start=3):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = 12

    wb.save(output_file)
    print(f"  Saved: {output_file}")


def create_comparison_excel(scores: List[Dict], output_file: str):
    """
    Create a model comparison sheet showing average scores by level.

    Args:
        scores: List of judge scores
        output_file: Output file path
    """
    print(f"\nCreating model comparison workbook: {output_file}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Model Comparison"

    # Organize scores
    model_scores = organize_by_model(scores)

    # Headers
    headers = ['Model', 'Level', 'Factuality', 'Helpfulness', 'Structure', 'Conciseness', 'Total Score']
    ws.append(headers)

    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Calculate averages for each model and level
    for model_id in sorted(model_scores.keys()):
        level_stats = defaultdict(lambda: {
            'factuality': [], 'helpfulness': [], 'structure': [],
            'conciseness': [], 'total': []
        })

        for score in model_scores[model_id]:
            level = score['level']
            level_stats[level]['factuality'].append(score['factuality'])
            level_stats[level]['helpfulness'].append(score['helpfulness'])
            level_stats[level]['structure'].append(score['structure'])
            level_stats[level]['conciseness'].append(score['conciseness'])
            level_stats[level]['total'].append(score['total_score'])

        # Write averages for each level
        for level in ['L0', 'L1', 'L2', 'L3', 'P']:
            if level in level_stats and level_stats[level]['total']:
                avg_fact = sum(level_stats[level]['factuality']) / len(level_stats[level]['factuality'])
                avg_help = sum(level_stats[level]['helpfulness']) / len(level_stats[level]['helpfulness'])
                avg_struct = sum(level_stats[level]['structure']) / len(level_stats[level]['structure'])
                avg_conc = sum(level_stats[level]['conciseness']) / len(level_stats[level]['conciseness'])
                avg_total = sum(level_stats[level]['total']) / len(level_stats[level]['total'])

                ws.append([
                    model_id,
                    level,
                    round(avg_fact, 2),
                    round(avg_help, 2),
                    round(avg_struct, 2),
                    round(avg_conc, 2),
                    round(avg_total, 2)
                ])

    # Format columns
    for i in range(1, 8):
        ws.column_dimensions[get_column_letter(i)].width = 15

    wb.save(output_file)
    print(f"  Saved: {output_file}")


def aggregate_results(runs_file: str, scores_file: str, output_dir: str):
    """
    Aggregate all results into Excel workbooks.

    Args:
        runs_file: Path to runs_llamacpp.jsonl
        scores_file: Path to scores_absolute.csv
        output_dir: Output directory for Excel files
    """
    print("="*80)
    print("LOADING DATA")
    print("="*80)

    # Load data
    results = load_inference_results(runs_file)
    scores = load_judge_scores(scores_file)

    # Create output directory
    os.makedirs(output_dir, exist_ok=True)

    print("\n" + "="*80)
    print("CREATING EXCEL WORKBOOKS")
    print("="*80)

    # Create quality scores workbook (detailed)
    quality_file = os.path.join(output_dir, 'quality_scores_detailed.xlsx')
    create_quality_excel(scores, quality_file)

    # Create model comparison workbook (summary)
    comparison_file = os.path.join(output_dir, 'model_comparison.xlsx')
    create_comparison_excel(scores, comparison_file)

    # Create energy workbook
    energy_file = os.path.join(output_dir, 'energy_per_run.xlsx')
    create_metric_excel(results, 'Energy', 'energy_joule', energy_file)

    # Create latency workbook
    latency_file = os.path.join(output_dir, 'latency_per_run.xlsx')
    create_metric_excel(results, 'Latency', 'latency_ms', latency_file)

    # Create output tokens workbook
    tokens_file = os.path.join(output_dir, 'output_tokens_per_run.xlsx')
    create_metric_excel(results, 'Output Tokens', 'completion_tokens', tokens_file)

    print("\n" + "="*80)
    print("AGGREGATION COMPLETE")
    print("="*80)
    print(f"\nGenerated files in {output_dir}:")
    print(f"  - quality_scores_detailed.xlsx  (per-question scores with reasoning)")
    print(f"  - model_comparison.xlsx          (average scores by model & level)")
    print(f"  - energy_per_run.xlsx")
    print(f"  - latency_per_run.xlsx")
    print(f"  - output_tokens_per_run.xlsx")


def main():
    parser = argparse.ArgumentParser(
        description="Aggregate benchmark results with absolute scoring into Excel workbooks"
    )

    parser.add_argument(
        '--runs',
        required=True,
        help='Path to runs_llamacpp.jsonl'
    )

    parser.add_argument(
        '--scores',
        required=True,
        help='Path to scores_absolute.csv'
    )

    parser.add_argument(
        '--outdir',
        required=True,
        help='Output directory for Excel files'
    )

    args = parser.parse_args()

    # Validate inputs
    if not os.path.exists(args.runs):
        print(f"Error: Runs file not found: {args.runs}")
        exit(1)

    if not os.path.exists(args.scores):
        print(f"Error: Scores file not found: {args.scores}")
        exit(1)

    # Run aggregation
    aggregate_results(args.runs, args.scores, args.outdir)


if __name__ == "__main__":
    main()
