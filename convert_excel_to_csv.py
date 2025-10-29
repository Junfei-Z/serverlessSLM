"""
Convert Generated Prompts Excel to CSV Format
Converts generated_prompts.xlsx (Sheet 1) to category-specific CSV files.
"""

import argparse
import os
import csv
from openpyxl import load_workbook
from collections import defaultdict


def convert_excel_to_csv(excel_file: str, output_dir: str, split_by_category: bool = True):
    """
    Convert prompts Excel to CSV files.

    Args:
        excel_file: Path to generated_prompts.xlsx
        output_dir: Output directory for CSV files
        split_by_category: If True, create separate CSV per category
    """
    print(f"Loading Excel file: {excel_file}")

    # Load workbook
    wb = load_workbook(excel_file, read_only=True)
    ws = wb['Prompts']

    # Read all data
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        print("Error: Empty worksheet")
        return

    # Get headers
    headers = rows[0]
    data_rows = rows[1:]

    print(f"Found {len(data_rows)} prompt tasks")

    # Create output directory
    os.makedirs(output_dir, exist_ok=True)

    if split_by_category:
        # Group by category
        category_data = defaultdict(list)

        for row in data_rows:
            row_dict = dict(zip(headers, row))
            category = row_dict.get('category', 'unknown').lower()
            category_data[category].append(row_dict)

        # Write separate CSV per category
        for category, cat_rows in category_data.items():
            output_file = os.path.join(output_dir, f'{category}_prompts_hierarchical.csv')

            with open(output_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(cat_rows)

            print(f"  Created: {output_file} ({len(cat_rows)} tasks)")

    else:
        # Write single CSV with all data
        output_file = os.path.join(output_dir, 'all_prompts_hierarchical.csv')

        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(data_rows)

        print(f"  Created: {output_file} ({len(data_rows)} tasks)")

    wb.close()
    print("Conversion complete!")


def main():
    parser = argparse.ArgumentParser(
        description="Convert generated_prompts.xlsx to CSV format"
    )

    parser.add_argument(
        '--excel',
        default='generated_prompts.xlsx',
        help='Path to generated_prompts.xlsx (default: generated_prompts.xlsx)'
    )

    parser.add_argument(
        '--outdir',
        default='data',
        help='Output directory for CSV files (default: data)'
    )

    parser.add_argument(
        '--split',
        action='store_true',
        default=True,
        help='Split into separate files per category (default: True)'
    )

    parser.add_argument(
        '--no-split',
        dest='split',
        action='store_false',
        help='Create single CSV file with all categories'
    )

    args = parser.parse_args()

    if not os.path.exists(args.excel):
        print(f"Error: Excel file not found: {args.excel}")
        exit(1)

    convert_excel_to_csv(args.excel, args.outdir, args.split)


if __name__ == "__main__":
    main()
