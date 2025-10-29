"""
Prompt Generation Script
Generates hierarchical prompts (L0, L1, L2, L3, P) from question.jsonl
and exports to Excel with prompt and token count sheets.
"""

import json
import tiktoken
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from typing import Dict, List, Tuple


class PromptGenerator:
    def __init__(self, model_name="gpt-4"):
        """Initialize with tiktoken encoder for accurate token counting."""
        try:
            self.encoder = tiktoken.encoding_for_model(model_name)
        except:
            self.encoder = tiktoken.get_encoding("cl100k_base")

    def count_tokens(self, text: str) -> int:
        """Count tokens in text using tiktoken."""
        return len(self.encoder.encode(text))

    def generate_l0(self, question: str) -> str:
        """L0: Base Prompt - The original task instruction."""
        return question

    def generate_l1(self, l0: str, category: str) -> str:
        """L1: Clarified Prompt - Adds clarity, scoring dimensions, basic constraints."""

        # Category-specific enhancements
        enhancements = {
            "writing": "\n\nPlease ensure your response includes:\n- A clear introduction, body, and conclusion\n- Coherent structure and logical flow\n- Appropriate tone and style for the context\n- Proper grammar and punctuation",

            "roleplay": "\n\nPlease ensure your response:\n- Stays in character throughout\n- Uses appropriate tone and mannerisms\n- Demonstrates understanding of the character's background and motivations\n- Maintains consistency with the character's known personality traits",

            "reasoning": "\n\nPlease ensure your response:\n- Clearly explains the logical reasoning process\n- Identifies key assumptions and constraints\n- Shows step-by-step deduction\n- Arrives at a well-justified conclusion",

            "math": "\n\nPlease ensure your response:\n- Shows all calculation steps clearly\n- Explains the mathematical reasoning\n- Provides the final answer in the appropriate format\n- Verifies the solution makes sense in context",

            "coding": "\n\nPlease ensure your response:\n- Includes clean, well-commented code\n- Follows best practices and coding standards\n- Handles edge cases appropriately\n- Provides explanation of the approach used",

            "extraction": "\n\nPlease ensure your response:\n- Accurately extracts all requested information\n- Follows the specified output format precisely\n- Organizes data in a clear and structured manner\n- Maintains consistency in formatting",

            "stem": "\n\nPlease ensure your response:\n- Uses accurate scientific terminology\n- Explains concepts clearly and thoroughly\n- Provides relevant examples or applications\n- Cites fundamental principles where appropriate",

            "humanities": "\n\nPlease ensure your response:\n- Demonstrates depth of understanding\n- Considers multiple perspectives where relevant\n- Supports arguments with evidence or examples\n- Maintains academic rigor and clarity"
        }

        enhancement = enhancements.get(category.lower(), "\n\nPlease ensure your response is clear, well-structured, and addresses all aspects of the question.")

        return l0 + enhancement

    def generate_l2(self, l1: str, category: str) -> str:
        """L2: Guided Prompt - Adds step-by-step execution instructions."""

        # Category-specific step-by-step guidance
        guidance = {
            "writing": "\n\nFollow this structured approach:\nStep 1: Plan your content - Outline the main points and supporting details\nStep 2: Craft the introduction - Hook the reader and present your thesis\nStep 3: Develop the body - Elaborate on each main point with evidence and examples\nStep 4: Write the conclusion - Summarize key points and provide closure\nStep 5: Review and refine - Check for coherence, clarity, and correctness",

            "roleplay": "\n\nFollow this roleplay framework:\nStep 1: Establish the character's mindset - Consider their current situation and emotional state\nStep 2: Respond authentically - Use language and expressions characteristic of the character\nStep 3: Incorporate character knowledge - Reference relevant background information naturally\nStep 4: Maintain consistency - Ensure actions and statements align with character traits\nStep 5: Engage meaningfully - Address the situation while staying true to character",

            "reasoning": "\n\nFollow this logical reasoning process:\nStep 1: Identify the problem - Clearly state what needs to be determined\nStep 2: Gather relevant information - List all given facts and constraints\nStep 3: Analyze relationships - Examine how different elements connect\nStep 4: Apply logical deduction - Work through the reasoning systematically\nStep 5: Validate the conclusion - Verify the answer against the original problem",

            "math": "\n\nFollow this problem-solving approach:\nStep 1: Understand the problem - Identify what is being asked and what is given\nStep 2: Plan the solution - Determine which mathematical concepts and formulas to use\nStep 3: Execute calculations - Work through the math step-by-step\nStep 4: Interpret results - Translate mathematical answer to the context\nStep 5: Verify solution - Check if the answer is reasonable and accurate",

            "coding": "\n\nFollow this development workflow:\nStep 1: Understand requirements - Clarify what the code needs to accomplish\nStep 2: Design the solution - Plan the algorithm and data structures\nStep 3: Implement the code - Write clean, modular code with appropriate comments\nStep 4: Test edge cases - Verify the code handles various scenarios correctly\nStep 5: Optimize if needed - Review for efficiency and readability improvements",

            "extraction": "\n\nFollow this extraction process:\nStep 1: Read carefully - Review the source material thoroughly\nStep 2: Identify target information - Locate all instances of requested data\nStep 3: Extract systematically - Gather each piece of information accurately\nStep 4: Format appropriately - Organize data according to specified structure\nStep 5: Verify completeness - Ensure all required information is included",

            "stem": "\n\nFollow this explanatory framework:\nStep 1: Define key terms - Explain relevant scientific concepts and terminology\nStep 2: Present core principles - Describe the fundamental laws or theories involved\nStep 3: Explain mechanisms - Detail how processes or phenomena occur\nStep 4: Provide examples - Illustrate with concrete instances or applications\nStep 5: Synthesize understanding - Connect concepts to the broader context",

            "humanities": "\n\nFollow this analytical approach:\nStep 1: Establish context - Provide relevant background information\nStep 2: Present main arguments - Articulate key positions or perspectives\nStep 3: Support with evidence - Use examples, data, or authoritative sources\nStep 4: Address counterpoints - Consider alternative viewpoints where appropriate\nStep 5: Draw conclusions - Synthesize insights and implications"
        }

        step_guidance = guidance.get(category.lower(), "\n\nFollow this systematic approach:\nStep 1: Analyze the question thoroughly\nStep 2: Organize your thoughts and main points\nStep 3: Develop your response systematically\nStep 4: Ensure all requirements are addressed\nStep 5: Review for completeness and accuracy")

        return l1 + step_guidance

    def generate_l3(self, l2: str, category: str) -> str:
        """L3: Few-Shot/Example-Guided Prompt - Adds examples and stylistic expectations."""

        # Category-specific examples and stylistic guidance
        examples = {
            "writing": "\n\nExample of effective structure:\nIntroduction: \"Literature serves as a mirror to society, reflecting both our greatest achievements and deepest struggles. Through the written word, we gain insights that transcend time and culture.\"\nBody: Develop each point with specific examples and thoughtful analysis, maintaining a consistent voice.\nConclusion: Synthesize the main ideas while leaving the reader with a lasting impression.\n\nStylistic expectations: Use varied sentence structure, incorporate transitions smoothly, and maintain an engaging yet appropriate tone throughout.",

            "roleplay": "\n\nExample of staying in character:\nIf roleplaying Sherlock Holmes: \"Elementary, my dear Watson. The mud on the suspect's shoes indicates a recent visit to the botanical gardens - note the distinctive reddish clay found only in that location. Combined with the pollen samples, we can deduce...\"\n\nStylistic expectations: Mirror the character's speech patterns, incorporate their typical mannerisms, reference their background knowledge naturally, and maintain their characteristic perspective throughout the interaction.",

            "reasoning": "\n\nExample of clear logical reasoning:\n\"Let's examine this systematically. We know that A is true, and we know that if A is true, then B must also be true. Therefore, B is true. Now, given that B is true and C is false, we can deduce that D cannot be the case because...\"\n\nStylistic expectations: Make your reasoning explicit and transparent. Use clear logical connectors (\"therefore,\" \"because,\" \"given that\") and ensure each step follows necessarily from the previous ones.",

            "math": "\n\nExample of clear mathematical reasoning:\n\"Given: The area of a rectangle is 48 square meters, and its length is 8 meters.\nSolution: Using the formula Area = length × width\n48 = 8 × width\nwidth = 48 ÷ 8 = 6 meters\nTherefore, the width is 6 meters.\"\n\nStylistic expectations: Show all work, label steps clearly, use proper mathematical notation, and explain the reasoning behind each operation.",

            "coding": "\n\nExample of well-structured code:\n```python\ndef find_max(numbers):\n    \"\"\"Find the maximum value in a list of numbers.\n    \n    Args:\n        numbers: List of numeric values\n    Returns:\n        The maximum value, or None if list is empty\n    \"\"\"\n    if not numbers:\n        return None\n    return max(numbers)\n```\n\nStylistic expectations: Include docstrings, use descriptive variable names, handle edge cases, write modular functions, and add comments for complex logic.",

            "extraction": "\n\nExample of proper extraction format:\nIf extracting from: \"John Smith, CEO of TechCorp, announced in New York...\"\nOutput: {\"person\": \"John Smith\", \"title\": \"CEO\", \"organization\": \"TechCorp\", \"location\": \"New York\"}\n\nStylistic expectations: Maintain consistent formatting, preserve exact values from source material, organize data logically, and ensure output is machine-readable if JSON/CSV format is required.",

            "stem": "\n\nExample of clear scientific explanation:\n\"Photosynthesis occurs in two main stages. In the light-dependent reactions, which take place in the thylakoid membranes, chlorophyll absorbs light energy to produce ATP and NADPH. These energy carriers are then used in the light-independent reactions (Calvin cycle) in the stroma to convert CO₂ into glucose.\"\n\nStylistic expectations: Use precise scientific terminology, explain processes mechanistically, connect concepts to real-world applications, and maintain accuracy while ensuring clarity.",

            "humanities": "\n\nExample of analytical writing:\n\"The Renaissance marked a pivotal shift in Western thought. For instance, the emergence of humanism challenged medieval scholasticism by emphasizing individual potential and classical learning. This is evidenced by works such as Pico della Mirandola's 'Oration on the Dignity of Man,' which argued that humans possess the capacity for self-determination.\"\n\nStylistic expectations: Balance breadth and depth, support claims with specific evidence, acknowledge complexity and multiple perspectives, and maintain scholarly rigor while remaining accessible."
        }

        example_guidance = examples.get(category.lower(), "\n\nExample of quality response:\nProvide specific, detailed information that directly addresses the question. Use concrete examples to illustrate abstract concepts. Maintain coherence and logical flow throughout.\n\nStylistic expectations: Demonstrate thoroughness, accuracy, and clear communication. Ensure your response is well-organized and easy to follow.")

        return l2 + example_guidance

    def generate_p(self, l3: str) -> str:
        """P: Placebo Prompt - Adds uninformative fluff without real content."""

        placebo_text = """

Please approach this task with the utmost care and attention to detail. I would greatly appreciate it if you could do your absolute best work on this request. Take all the time you need to ensure that every possible aspect is thoroughly considered and addressed with the highest level of diligence and thoughtfulness.

It is of paramount importance that your response be as comprehensive, detailed, and well-considered as humanly possible. Please ensure that you explore every conceivable angle and perspective with exceptional thoroughness and meticulous precision. Your careful attention to even the smallest details will be immensely valued and appreciated.

I have complete confidence in your abilities and trust that you will deliver an outstanding response that exemplifies the very best of your capabilities. Thank you so much in advance for your dedication and commitment to excellence in completing this task."""

        return l3 + placebo_text

    def generate_all_levels(self, question_data: Dict) -> Dict[str, str]:
        """Generate all prompt levels for a single question."""

        # Use the first turn as the base question
        base_question = question_data['turns'][0]
        category = question_data['category']

        # Generate each level
        l0 = self.generate_l0(base_question)
        l1 = self.generate_l1(l0, category)
        l2 = self.generate_l2(l1, category)
        l3 = self.generate_l3(l2, category)
        p = self.generate_p(l3)

        return {
            'L0': l0,
            'L1': l1,
            'L2': l2,
            'L3': l3,
            'P': p
        }

    def get_token_counts(self, prompts: Dict[str, str]) -> Dict[str, int]:
        """Calculate token counts for all prompt levels."""
        return {
            f'{level}_tokens': self.count_tokens(text)
            for level, text in prompts.items()
        }


def read_jsonl(filepath: str) -> List[Dict]:
    """Read JSONL file and return list of question objects."""
    questions = []
    with open(filepath, 'r', encoding='utf-8') as f:
        for line in f:
            if line.strip():
                questions.append(json.loads(line))
    return questions


def create_excel_output(questions: List[Dict], prompts_data: List[Tuple], output_file: str):
    """Create Excel workbook with two sheets: Prompts and Token Counts."""

    wb = Workbook()

    # Sheet 1: Prompts
    ws1 = wb.active
    ws1.title = "Prompts"

    # Headers for Sheet 1
    headers1 = ['question_id', 'category', 'topic', 'L0', 'L1', 'L2', 'L3', 'P']
    ws1.append(headers1)

    # Style headers
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add data rows
    for q, prompts in prompts_data:
        # Use first turn as topic (truncated if too long)
        topic = q['turns'][0][:50] + '...' if len(q['turns'][0]) > 50 else q['turns'][0]

        row = [
            q['question_id'],
            q['category'],
            topic,
            prompts['L0'],
            prompts['L1'],
            prompts['L2'],
            prompts['L3'],
            prompts['P']
        ]
        ws1.append(row)

    # Adjust column widths for Sheet 1
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 30
    for col in ['D', 'E', 'F', 'G', 'H']:
        ws1.column_dimensions[col].width = 50

    # Enable text wrapping for prompt columns
    for row in ws1.iter_rows(min_row=2):
        for cell in row[3:]:  # Columns D through H (prompts)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Sheet 2: Token Counts
    ws2 = wb.create_sheet("Token Counts")

    # Headers for Sheet 2
    headers2 = ['question_id', 'category', 'L0_tokens', 'L1_tokens', 'L2_tokens', 'L3_tokens', 'P_tokens']
    ws2.append(headers2)

    # Style headers
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add token count data
    generator = PromptGenerator()
    for q, prompts in prompts_data:
        token_counts = generator.get_token_counts(prompts)

        row = [
            q['question_id'],
            q['category'],
            token_counts['L0_tokens'],
            token_counts['L1_tokens'],
            token_counts['L2_tokens'],
            token_counts['L3_tokens'],
            token_counts['P_tokens']
        ]
        ws2.append(row)

    # Adjust column widths for Sheet 2
    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 15

    # Save workbook
    wb.save(output_file)
    print(f"Excel file saved: {output_file}")


def main():
    """Main execution function."""
    print("Starting prompt generation...")

    # Configuration
    input_file = "question.jsonl"
    output_file = "generated_prompts.xlsx"

    # Read questions
    print(f"Reading questions from {input_file}...")
    questions = read_jsonl(input_file)
    print(f"Loaded {len(questions)} questions")

    # Generate prompts
    print("Generating prompts for all levels...")
    generator = PromptGenerator()
    prompts_data = []

    for i, question in enumerate(questions, 1):
        print(f"Processing question {i}/{len(questions)} (ID: {question['question_id']})")
        prompts = generator.generate_all_levels(question)
        prompts_data.append((question, prompts))

        # Show token counts for verification
        token_counts = generator.get_token_counts(prompts)
        print(f"  Tokens - L0: {token_counts['L0_tokens']}, "
              f"L1: {token_counts['L1_tokens']}, "
              f"L2: {token_counts['L2_tokens']}, "
              f"L3: {token_counts['L3_tokens']}, "
              f"P: {token_counts['P_tokens']}")

        # Verify monotonic increase
        tokens = [token_counts['L0_tokens'], token_counts['L1_tokens'],
                 token_counts['L2_tokens'], token_counts['L3_tokens'],
                 token_counts['P_tokens']]
        if tokens != sorted(tokens):
            print(f"  WARNING: Token counts not monotonically increasing!")

    # Create Excel output
    print(f"\nCreating Excel output...")
    create_excel_output(questions, prompts_data, output_file)

    print("\n✓ Prompt generation complete!")
    print(f"✓ Output saved to: {output_file}")
    print(f"✓ Total questions processed: {len(questions)}")


if __name__ == "__main__":
    main()
